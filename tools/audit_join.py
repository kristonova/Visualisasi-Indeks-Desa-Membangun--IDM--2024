#!/usr/bin/env python3
r"""Audit the IDM 2024 Excel -> compact JSON -> GeoJSON-code join.

The dashboard joins records only by the 10-digit Ministry of Home Affairs code.
This command reproduces that exact join, distinguishes villages from urban wards,
and writes a deterministic report that can be reviewed or checked in CI.

Run from the repository root:

    .venv\Scripts\python tools\audit_join.py
    .venv\Scripts\python tools\audit_join.py --strict
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - exercised only on a broken setup
    raise SystemExit(
        "openpyxl belum terpasang. Jalankan audit dengan .venv\\Scripts\\python "
        "atau pasang openpyxl terlebih dahulu."
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "data" / "indeks-desa-membangun-tahun-2024-hasil-pemutakhiran.xlsx"
DEFAULT_IDM_DIR = ROOT / "data" / "idm" / "prov"
DEFAULT_GEO_DIR = ROOT / "data" / "geo" / "desa"
DEFAULT_REPORT = ROOT / "data" / "geo" / "join-audit.json"

STATUS_TO_CODE = {
    None: 0,
    "": 0,
    "MANDIRI": 1,
    "MAJU": 2,
    "BERKEMBANG": 3,
    "TERTINGGAL": 4,
    "SANGAT TERTINGGAL": 5,
}
SOURCE_TO_CODE = {None: 0, "": 0, "Update 2023": 1, "Server PDN": 2}
UNIT_KIND = {"1": "kelurahan", "2": "desa", "3": "desa_adat"}


def digits(value: Any) -> str:
    """Return only decimal digits without converting large codes through float."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\D", "", str(value))


def unit_kind(code: str) -> str:
    return UNIT_KIND.get(code[6:7], "lainnya")


def normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii").upper()
    return re.sub(r"[^A-Z0-9]", "", text)


def numeric(value: Any) -> Optional[float]:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    value = float(value)
    return value if math.isfinite(value) else None


def same_number(left: Any, right: Any, tolerance: float = 0.00005) -> bool:
    left_num, right_num = numeric(left), numeric(right)
    if left_num is None or right_num is None:
        return left_num is None and right_num is None
    return abs(left_num - right_num) <= tolerance


def duplicate_codes(records: List[Dict[str, Any]]) -> List[str]:
    counts = Counter(record["code"] for record in records if record["code"])
    return sorted(code for code, count in counts.items() if count > 1)


def load_excel(path: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "IDM 2024" not in workbook.sheetnames:
        raise SystemExit(f"Sheet 'IDM 2024' tidak ditemukan di {path}")
    sheet = workbook["IDM 2024"]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    columns = {str(name).strip(): index for index, name in enumerate(header) if name is not None}
    required = [
        "KODE_PROV",
        "NAMA_PROVINSI",
        "KODE_KAB",
        "NAMA_KABUPATEN",
        "KODE_KEC",
        "NAMA_KECAMATAN",
        "KODE_DESA",
        "NAMA_DESA",
        "IKS_2024",
        "IKE_2024",
        "IKL_2024",
        "NILAI_IDM_2024",
        "STATUS_IDM_2024",
        "Keterangan",
    ]
    missing_columns = [name for name in required if name not in columns]
    if missing_columns:
        raise SystemExit("Kolom Excel tidak lengkap: " + ", ".join(missing_columns))

    records: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        get = lambda name: row[columns[name]] if columns[name] < len(row) else None
        code = digits(get("KODE_DESA"))
        source_text = get("Keterangan")
        status_text = get("STATUS_IDM_2024")
        record = {
            "code": code,
            "province_code": digits(get("KODE_PROV")).zfill(2),
            "province": str(get("NAMA_PROVINSI") or "").strip(),
            "regency_code": digits(get("KODE_KAB")).zfill(4),
            "regency": str(get("NAMA_KABUPATEN") or "").strip(),
            "district_code": digits(get("KODE_KEC")).zfill(6),
            "district": str(get("NAMA_KECAMATAN") or "").strip(),
            "name": str(get("NAMA_DESA") or "").strip(),
            "iks": numeric(get("IKS_2024")),
            "ike": numeric(get("IKE_2024")),
            "ikl": numeric(get("IKL_2024")),
            "idm": numeric(get("NILAI_IDM_2024")),
            "status": STATUS_TO_CODE.get(status_text, -1),
            "source": SOURCE_TO_CODE.get(source_text, -1),
            "source_label": str(source_text or "Reguler"),
            "source_note": (
                str(get("IKS_2024")).strip()
                if get("IKS_2024") not in (None, "") and numeric(get("IKS_2024")) is None
                else None
            ),
            "row": row_number,
        }
        if len(code) != 10 or not code.isdigit():
            invalid.append({"row": row_number, "code": code, "name": record["name"]})
        records.append(record)
    workbook.close()
    return records, invalid


def load_idm_json(directory: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    records: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    for path in sorted(directory.glob("*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        for district in payload.get("kec", []):
            district_code = str(district.get("k", ""))
            for village in district.get("ds", []):
                code = district_code + str(village[0])
                record = {
                    "code": code,
                    "province_code": code[:2],
                    "regency_code": str(district.get("kb", "")),
                    "district_code": district_code,
                    "district": str(district.get("nm", "")),
                    "name": str(village[1] or ""),
                    "idm": village[2],
                    "iks": village[3],
                    "ike": village[4],
                    "ikl": village[5],
                    "status": village[6],
                    "source": village[7],
                    "file": path.name,
                }
                if len(code) != 10 or not code.isdigit():
                    invalid.append({"file": path.name, "code": code, "name": record["name"]})
                records.append(record)
    return records, invalid


def load_geo_json(directory: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    records: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    for path in sorted(directory.glob("*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        for feature in payload.get("f", []):
            code = str(feature.get("k", ""))
            record = {
                "code": code,
                "province_code": code[:2],
                "name": str(feature.get("nm", "")),
                "kind": unit_kind(code),
                "file": path.name,
            }
            if len(code) != 10 or not code.isdigit():
                invalid.append({"file": path.name, "code": code, "name": record["name"]})
            records.append(record)
    return records, invalid


def compact_record(record: Dict[str, Any], include_note: bool = False) -> Dict[str, Any]:
    output = {
        "code": record["code"],
        "name": record.get("name", "") or "(tanpa nama)",
        "province_code": record["code"][:2],
        "regency_code": record["code"][:4],
        "district_code": record["code"][:6],
    }
    if record.get("district"):
        output["district"] = record["district"]
    if record.get("regency"):
        output["regency"] = record["regency"]
    if record.get("province"):
        output["province"] = record["province"]
    if include_note and record.get("source_note"):
        output["reason"] = record["source_note"]
    if include_note and record.get("source_label"):
        output["source"] = record["source_label"]
    return output


def compare_excel_and_json(
    excel_by_code: Dict[str, Dict[str, Any]],
    idm_by_code: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    mismatches: List[Dict[str, Any]] = []
    fields = ("name", "district_code", "regency_code", "idm", "iks", "ike", "ikl", "status", "source")
    for code in sorted(excel_by_code.keys() & idm_by_code.keys()):
        excel_record, json_record = excel_by_code[code], idm_by_code[code]
        changed: Dict[str, Dict[str, Any]] = {}
        for field in fields:
            left, right = excel_record.get(field), json_record.get(field)
            same = same_number(left, right) if field in {"idm", "iks", "ike", "ikl"} else left == right
            if not same:
                changed[field] = {"excel": left, "json": right}
        if changed:
            mismatches.append({"code": code, "fields": changed})
    return mismatches


def province_summary(
    excel_codes: Set[str], geo_codes: Set[str], no_score_codes: Set[str]
) -> List[Dict[str, Any]]:
    provinces = sorted({code[:2] for code in excel_codes | geo_codes})
    output = []
    for province in provinces:
        excel = {code for code in excel_codes if code.startswith(province)}
        geo = {code for code in geo_codes if code.startswith(province)}
        geo_only = geo - excel
        output.append(
            {
                "province_code": province,
                "idm_rows": len(excel),
                "geo_features": len(geo),
                "joined": len(excel & geo),
                "idm_without_geometry": len(excel - geo),
                "geo_kelurahan_outside_idm": sum(unit_kind(code) == "kelurahan" for code in geo_only),
                "unexpected_geo_village_without_idm": sum(unit_kind(code) != "kelurahan" for code in geo_only),
                "idm_without_score": sum(code.startswith(province) for code in no_score_codes),
            }
        )
    return output


def display_path(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def audit(args: argparse.Namespace) -> Tuple[Dict[str, Any], List[str]]:
    excel_records, excel_invalid = load_excel(args.excel)
    idm_records, idm_invalid = load_idm_json(args.idm_dir)
    geo_records, geo_invalid = load_geo_json(args.geo_dir)

    excel_duplicates = duplicate_codes(excel_records)
    idm_duplicates = duplicate_codes(idm_records)
    geo_duplicates = duplicate_codes(geo_records)
    excel_by_code = {record["code"]: record for record in excel_records}
    idm_by_code = {record["code"]: record for record in idm_records}
    geo_by_code = {record["code"]: record for record in geo_records}

    excel_codes, idm_codes, geo_codes = set(excel_by_code), set(idm_by_code), set(geo_by_code)
    joined = idm_codes & geo_codes
    idm_without_geo = idm_codes - geo_codes
    geo_without_idm = geo_codes - idm_codes
    geo_kelurahan = {code for code in geo_without_idm if unit_kind(code) == "kelurahan"}
    unexpected_geo = geo_without_idm - geo_kelurahan
    no_score = {
        code
        for code, record in excel_by_code.items()
        if record["idm"] is None or record["status"] == 0
    }
    metric_mismatches = compare_excel_and_json(excel_by_code, idm_by_code)
    joined_name_mismatches = []
    for code in sorted(joined):
        idm_name = excel_by_code[code].get("name", "")
        geo_name = geo_by_code[code].get("name", "")
        if normalize_name(idm_name) != normalize_name(geo_name):
            joined_name_mismatches.append(
                {
                    "code": code,
                    "idm_name": idm_name or "(kosong)",
                    "geo_name": geo_name or "(kosong)",
                    "kind": "missing_idm_name" if not idm_name else "name_difference",
                }
            )

    report = {
        "schema_version": 1,
        "sources": {
            "excel": display_path(args.excel),
            "idm_json": display_path(args.idm_dir),
            "geo_json": display_path(args.geo_dir),
        },
        "summary": {
            "excel_rows": len(excel_records),
            "idm_json_rows": len(idm_records),
            "geo_features": len(geo_records),
            "joined_idm_to_geometry": len(joined),
            "idm_without_geometry": len(idm_without_geo),
            "geo_kelurahan_outside_idm": len(geo_kelurahan),
            "unexpected_geo_village_without_idm": len(unexpected_geo),
            "idm_without_score": len(no_score),
            "joined_name_mismatches": len(joined_name_mismatches),
            "join_rate_percent": round(100 * len(joined) / max(1, len(idm_codes)), 4),
        },
        "integrity": {
            "excel_duplicate_codes": excel_duplicates,
            "idm_json_duplicate_codes": idm_duplicates,
            "geo_duplicate_codes": geo_duplicates,
            "excel_invalid_codes": excel_invalid,
            "idm_json_invalid_codes": idm_invalid,
            "geo_invalid_codes": geo_invalid,
            "excel_missing_from_idm_json": sorted(excel_codes - idm_codes),
            "idm_json_missing_from_excel": sorted(idm_codes - excel_codes),
            "excel_json_field_mismatches": metric_mismatches,
        },
        "by_province": province_summary(idm_codes, geo_codes, no_score),
        "idm_without_score": [
            compact_record(excel_by_code[code], include_note=True) for code in sorted(no_score)
        ],
        "joined_name_mismatches": joined_name_mismatches,
        "idm_without_geometry": [
            compact_record(excel_by_code[code]) for code in sorted(idm_without_geo)
        ],
        "unexpected_geo_village_without_idm": [
            dict(compact_record(geo_by_code[code]), kind=unit_kind(code))
            for code in sorted(unexpected_geo)
        ],
        "geo_kelurahan_outside_idm": {
            "count": len(geo_kelurahan),
            "explanation": "Kode unit 1xxx adalah kelurahan dan berada di luar cakupan IDM desa.",
            "sample": [compact_record(geo_by_code[code]) for code in sorted(geo_kelurahan)[:20]],
        },
    }

    failures: List[str] = []
    if excel_duplicates or idm_duplicates or geo_duplicates:
        failures.append("ditemukan kode duplikat")
    if excel_invalid or idm_invalid or geo_invalid:
        failures.append("ditemukan kode bukan 10 digit")
    if excel_codes != idm_codes or metric_mismatches:
        failures.append("JSON IDM tidak identik dengan sumber Excel")
    if unexpected_geo:
        failures.append("ada geometri desa/desa adat yang gagal join ke IDM")
    if len(idm_without_geo) > args.max_idm_without_geometry:
        failures.append(
            f"IDM tanpa geometri {len(idm_without_geo)} melebihi batas "
            f"{args.max_idm_without_geometry}"
        )
    if len(joined_name_mismatches) > args.max_joined_name_mismatches:
        failures.append(
            f"mismatch nama pada kode yang join {len(joined_name_mismatches)} melebihi batas "
            f"{args.max_joined_name_mismatches}"
        )
    return report, failures


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--idm-dir", type=Path, default=DEFAULT_IDM_DIR)
    parser.add_argument("--geo-dir", type=Path, default=DEFAULT_GEO_DIR)
    parser.add_argument("--out", type=Path, default=DEFAULT_REPORT)
    parser.add_argument(
        "--max-idm-without-geometry",
        type=int,
        default=335,
        help="fail in --strict mode if the known geometry gap grows (default: 335)",
    )
    parser.add_argument(
        "--max-joined-name-mismatches",
        type=int,
        default=3,
        help="fail in --strict mode if exact-code name mismatches grow (default: 3)",
    )
    parser.add_argument("--strict", action="store_true", help="return a non-zero exit code on integrity failures")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    for attr in ("excel", "idm_dir", "geo_dir", "out"):
        value = getattr(args, attr)
        if not value.is_absolute():
            setattr(args, attr, (ROOT / value).resolve())

    report, failures = audit(args)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    summary = report["summary"]
    print(f"Excel / JSON IDM : {summary['excel_rows']:,} / {summary['idm_json_rows']:,}")
    print(f"Fitur geometri   : {summary['geo_features']:,}")
    print(f"Join tepat       : {summary['joined_idm_to_geometry']:,} ({summary['join_rate_percent']:.4f}%)")
    print(f"IDM tanpa geometri: {summary['idm_without_geometry']:,}")
    print(f"Kelurahan di luar IDM: {summary['geo_kelurahan_outside_idm']:,}")
    print(f"Desa geo gagal join: {summary['unexpected_geo_village_without_idm']:,}")
    print(f"Desa tanpa skor : {summary['idm_without_score']:,}")
    print(f"Nama beda pada kode sama: {summary['joined_name_mismatches']:,}")
    print(f"Laporan          : {args.out}")
    if failures:
        print("Masalah integritas: " + "; ".join(failures), file=sys.stderr)
    return 1 if args.strict and failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
